import datetime
import os

import openpyxl
from lxml import etree
import requests
import xlwt
from tkinter import simpledialog
import tkinter
import xlrd
from xlutils.copy import copy

participant_name = []
shareholding = []
sharehoding_percent = []  # 差额
total=[]#总数
issued_share=[]#已发行股票总数


participant_name1 = []
shareholding1 = []
sharehoding_percent1 = []  # 差额
total1=[]#第二次爬取的总数
issued_share1=[]#第二次已发行股票总数

# hkex披露易1529定向爬虫
class HKex_Search:


    def __init__(self):
        self.URL  = "https://www3.hkexnews.hk/sdw/search/searchsdw_c.aspx"
        self.path = 'D:/披露易每日定向数据/'



    def get_stockdata(self, searchDate, stockcode):

        ''':argument:对象参数包括"20XX/XX/XX“和股票代号'''

        header = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9'
        }
        postdata = {
            '__EVENTTARGET': 'btnSearch',
            '__EVENTARGUMENT': '',
            '__VIEWSTATE': "/wEPDwUKMTY0ODYwNTA0OWRkM79k2SfZ+VkDy88JRhbk+XZIdM0=",  # 表单数据提交
            "__VIEWSTATEGENERATOR": "3B50BBBD",
            "sortBy": "shareholding",
            "sortDirection": "desc",
            "txtShareholdingDate": searchDate,
            "txtStockCode": stockcode,
        }

        response = requests.post(url=self.URL, data=postdata, headers=header)
        html0 = response.content.decode("utf-8")  # 确认页面爬取内容
        # print(html0)
        parse = etree.HTML(response.text)  # 转换为页面树
        # data_participant_name = parse.xpath('//div[@id="pnlResultNormal"]//table//td[@class="col-participant-name"]//div[@class="mobile-list-body"]/text()')#券商名
        # data_sharehoding=parse.xpath('//div[@id="pnlResultNormal"]//table//td[@class="col-shareholding text-right"]//div[@class="mobile-list-body"]/text()')#持股量
        # data_sharehoding_percent=parse.xpath('//div[@id="pnlResultNormal"]//table//td[@class="col-shareholding-percent text-right"]//div[@class="mobile-list-body"]/text()')#占已发行股份/权证/单位百分比
        # print(data_participant_name,data_sharehoding,data_sharehoding_percent)#验证输出

        for i in range(len(parse.xpath(
                '//div[@id="pnlResultNormal"]//table//td[@class="col-participant-name"]//div[@class="mobile-list-body"]/text()'))):
            participant_name.append(parse.xpath(
                '//div[@id="pnlResultNormal"]//table//td[@class="col-participant-name"]//div[@class="mobile-list-body"]/text()')[i])
            shareholding.append(parse.xpath(
                '//div[@id="pnlResultNormal"]//table//td[@class="col-shareholding text-right"]//div[@class="mobile-list-body"]/text()')[i])
            sharehoding_percent.append(parse.xpath(
                '//div[@id="pnlResultNormal"]//table//td[@class="col-shareholding-percent text-right"]//div[@class="mobile-list-body"]/text()')[i])  # 将每个元素放入列表

        total.append(parse.xpath('//div[@id="pnlResultSummary"]//div[@class="ccass-search-datarow ccass-search-total"]//div[@class="shareholding"]//div[@class="value"]/text()'))#持股总数
        issued_share.append((parse.xpath('//div[@id="pnlResultSummary"]//div[@class="summary-value"]/text()')))#已发行股份总数


    def get_stockdata1(self, searchDate, stockcode):

        ''':argument:对象参数包括"20XX/XX/XX“和股票代号'''

        header = {
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9'
        }
        postdata = {
            '__EVENTTARGET': 'btnSearch',
            '__EVENTARGUMENT': '',
            '__VIEWSTATE': "/wEPDwUKMTY0ODYwNTA0OWRkM79k2SfZ+VkDy88JRhbk+XZIdM0=",  # 表单数据提交
            "__VIEWSTATEGENERATOR": "3B50BBBD",
            "sortBy": "shareholding",
            "sortDirection": "desc",
            "txtShareholdingDate": searchDate,
            "txtStockCode": stockcode,
        }

        response = requests.post(url=self.URL, data=postdata, headers=header)
        html0 = response.content.decode("utf-8")  # 确认页面爬取内容
        # print(html0)
        parse = etree.HTML(response.text)  # 转换为页面树
        # data_participant_name = parse.xpath('//div[@id="pnlResultNormal"]//table//td[@class="col-participant-name"]//div[@class="mobile-list-body"]/text()')#券商名
        # data_sharehoding=parse.xpath('//div[@id="pnlResultNormal"]//table//td[@class="col-shareholding text-right"]//div[@class="mobile-list-body"]/text()')#持股量
        # data_sharehoding_percent=parse.xpath('//div[@id="pnlResultNormal"]//table//td[@class="col-shareholding-percent text-right"]//div[@class="mobile-list-body"]/text()')#占已发行股份/权证/单位百分比
        # print(data_participant_name,data_sharehoding,data_sharehoding_percent)#验证输出

        for i in range(len(parse.xpath(
                '//div[@id="pnlResultNormal"]//table//td[@class="col-participant-name"]//div[@class="mobile-list-body"]/text()'))):
            participant_name1.append(parse.xpath(
                '//div[@id="pnlResultNormal"]//table//td[@class="col-participant-name"]//div[@class="mobile-list-body"]/text()')[i])
            shareholding1.append(parse.xpath(
                '//div[@id="pnlResultNormal"]//table//td[@class="col-shareholding text-right"]//div[@class="mobile-list-body"]/text()')[i])
            sharehoding_percent1.append(parse.xpath(
                '//div[@id="pnlResultNormal"]//table//td[@class="col-shareholding-percent text-right"]//div[@class="mobile-list-body"]/text()')[i])  # 将每个元素放入列表

        total1.append(parse.xpath('//div[@id="pnlResultSummary"]//div[@class="ccass-search-datarow ccass-search-total"]//div[@class="shareholding"]//div[@class="value"]/text()'))#持股总数
        issued_share1.append((parse.xpath('//div[@id="pnlResultSummary"]//div[@class="summary-value"]/text()')))#已发行股份总数

    def getYesterday(self):
        today = datetime.date.today()
        oneday = datetime.timedelta(days=1)
        yesterday = today - oneday
        return yesterday

    def file_save(self, participant_name_final, shareholding_final, sharehoding_percent_final, input_stock_code,
                  input_date_data):  # 在D盘新建文件夹并将信息保存进excel

        input_date_data = str.replace(input_date_data, '/', '-')  # 替换‘/’否侧会导致路径解析错误
        if os.path.exists(self.path) == True:  # 判断路径存在
            if os.path.exists(self.path + '{}'.format(input_date_data) + '_' + input_stock_code + '.xlsx'):  # 判断文件存在
                os.system(
                    self.path + '{}'.format(input_date_data) + '_' + input_stock_code + '.xlsx')  # 自动打开存在路径下已有的excel
                return 0
            if os.path.exists(
                    self.path + '{}'.format(input_date_data) + '_' + input_stock_code + '.xlsx') == False:  # 误删文件后重新爬取下载
                workbook = xlwt.Workbook(encoding='utf-8')
                data_sheet = workbook.add_sheet("{}".format(input_date_data))
                data_sheet.write(0, 0, label='券商名')
                data_sheet.write(0, 1, label='持股量')
                data_sheet.write(0, 2, label='流通比')
                data_sheet.write(0,5,label='于中央结算系统的持股量总数')
                data_sheet.write(0,6,label='已发行股份/权证/单位')
                for i in range(len(participant_name_final)):
                    data_sheet.write(i + 1, 0, label=participant_name_final[i])#券商名
                    data_sheet.write(i + 1, 1, label=shareholding_final[i])#持股量
                    data_sheet.write(i + 1, 2, label=sharehoding_percent_final[i])#流通比

                data_sheet.write(1,5,label=total[0])
                data_sheet.write(1,6,label=issued_share[0])
                workbook.save(self.path + '{}'.format(input_date_data) + '_' + input_stock_code + '.xlsx')
                os.system(self.path + '{}'.format(input_date_data) + '_' + input_stock_code + '.xlsx')  # 打开文件
        elif os.path.exists(self.path) == False:
            os.makedirs(self.path)
            workbook = xlwt.Workbook(encoding='utf-8')
            data_sheet = workbook.add_sheet("{}".format(input_date_data))
            data_sheet.write(0, 0, label='券商名')
            data_sheet.write(0, 1, label='持股量')
            data_sheet.write(0, 2, label='流通比')
            data_sheet.write(0,5,label='于中央结算系统的持股量总数')
            data_sheet.write(0,6,label='已发行股份/权证/单位')
            for i in range(len(participant_name_final)):
                data_sheet.write(i + 1, 0, label=participant_name_final[i])
                data_sheet.write(i + 1, 1, label=shareholding_final[i])
                data_sheet.write(i + 1, 2, label=sharehoding_percent_final[i])

            data_sheet.write(1,5,label=total[0])#总数
            data_sheet.write(1,6,label=issued_share[0])#已发行股份
            workbook.save(self.path + '{}'.format(input_date_data) + '_' + input_stock_code + '.xlsx')
            os.system(self.path + '{}'.format(input_date_data) + '_' + input_stock_code + '.xlsx')  # 生成并打开excel



    def file_save1(self, participant_name_final, shareholding_final, sharehoding_percent_final, input_stock_code,
              input_date_start,input_date_final):  # 在D盘新建文件夹并将信息保存进excel
      input_date_start = str.replace(input_date_start, '/', '-')  # 替换‘/’否侧会导致路径解析错误
      input_date_final=str.replace(input_date_final,'/','-') #替换‘/’否侧会导致路径解析错误
      filename=self.path + '{}'.format(input_date_start) + '_'+input_date_final+ '_' + input_stock_code + '.xlsx'#已创建的文件路径名
      # sheet_name="{}".format(input_date_start+'-'+input_date_final)#表格名字

      if os.path.exists(self.path) == True:# 判断路径存在
        if os.path.exists(filename)==True:
          workbook=xlrd.open_workbook(filename)
          write_book=copy(workbook)
          write_book.get_sheet(0).write(0, 10, label='券商名')
          write_book.get_sheet(0).write(0,11,label='持股量')
          write_book.get_sheet(0).write(0, 12, label='流通比')

          write_book.get_sheet(0).write(0,15,label='于中央结算系统的持股量总数')
          write_book.get_sheet(0).write(0,16,label='已发行股份/权证/单位')
          for i in range(len(participant_name_final)):#循环添加数据
            write_book.get_sheet(0).write(i + 1, 10, label=participant_name_final[i])
            write_book.get_sheet(0).write(i + 1, 11, label=shareholding_final[i])
            write_book.get_sheet(0).write(i + 1, 12, label=sharehoding_percent_final[i])
          write_book.get_sheet(0).write(1,15,label=total[0])#总数
          write_book.get_sheet(0).write(1,16,label=issued_share[0])#已发行股份

          write_book.save(filename)


    def get_sheetdata(self,filename,sheetname):
        workbook=xlrd.open_workbook(filename)
        get_sheet=workbook.sheet_by_name(sheetname)

        return  sheetname#获取的数据


    def file_save2(self, participant_name_final, shareholding_final, sharehoding_percent_final, input_stock_code,
                   input_date_start,input_date_final):  # 在D盘新建文件夹并将信息保存进excel,比file多了一个形参input_data_final

        input_date_start = str.replace(input_date_start, '/', '-')  # 替换‘/’否侧会导致路径解析错误
        input_date_final=str.replace(input_date_final,'/','-')
        if os.path.exists(self.path) == True:  # 判断主目录"披露易文件爬取"路径存在
            if os.path.exists(self.path + '{}'.format(input_date_start) + '_'+input_date_final+'_' + input_stock_code + '.xlsx'):  # 判断文件存在
                os.system(
                    self.path + '{}'.format(input_date_start) + '_' +input_date_final+'_'+ input_stock_code + '.xlsx')  # 自动打开存在路径下已有的excel
                return 0
            if os.path.exists(
                    self.path + '{}'.format(input_date_start) + '_' +input_date_final+'_'+ input_stock_code + '.xlsx') == False:  # 文件不存在（误删） 重新爬取下载
                workbook = xlwt.Workbook(encoding='utf-8')
                data_sheet = workbook.add_sheet("{}".format(input_date_start+'-'+input_date_final))
                data_sheet.write(0, 0, label='券商名')
                data_sheet.write(0, 1, label='持股量')
                data_sheet.write(0, 2, label='流通比')
                data_sheet.write(0,5,label='于中央结算系统的持股量总数')
                data_sheet.write(0,6,label='已发行股份/权证/单位')
                for i in range(len(participant_name_final)):
                    data_sheet.write(i + 1, 0, label=participant_name_final[i])#券商名
                    data_sheet.write(i + 1, 1, label=shareholding_final[i])#持股量
                    data_sheet.write(i + 1, 2, label=sharehoding_percent_final[i])#流通比

                data_sheet.write(1,5,label=total[0])
                data_sheet.write(1,6,label=issued_share[0])
                workbook.save(self.path + '{}'.format(input_date_start) + '_'+input_date_final+'_' + input_stock_code + '.xlsx')

        elif os.path.exists(self.path) == False:#主目录路径不存在创建路径并添加数据
            os.makedirs(self.path)
            workbook = xlwt.Workbook(encoding='utf-8')
            data_sheet = workbook.add_sheet("{}".format(input_date_start)+'-'+input_date_final)
            data_sheet.write(0, 0, label='券商名')
            data_sheet.write(0, 1, label='持股量')
            data_sheet.write(0, 2, label='流通比')
            data_sheet.write(0,5,label='于中央结算系统的持股量总数')
            data_sheet.write(0,6,label='已发行股份/权证/单位')
            for i in range(len(participant_name_final)):
                data_sheet.write(i + 1, 0, label=participant_name_final[i])
                data_sheet.write(i + 1, 1, label=shareholding_final[i])
                data_sheet.write(i + 1, 2, label=sharehoding_percent_final[i])

            data_sheet.write(1,5,label=total[0])#总数
            data_sheet.write(1,6,label=issued_share[0])#已发行股份
            workbook.save(self.path + '{}'.format(input_date_start) + '_'+input_date_final+'_' + input_stock_code + '.xlsx')

    def main_window(self):  # 主窗口

        window = tkinter.Tk()
        button1 = tkinter.Button(window, text='获取指定日期的披露易持股信息',bg='Red',command=self.getData)  # 获取当天披露易持股信息
        button2 = tkinter.Button(window, text='获取指定日期之差的持股差额', bg='SkyBlue', command=self.getBalance)#获取两天日期的数据并计算差额
        button3=tkinter.Button(window,text='退出',bg='yellow',command=window.quit)
        button1.grid(row=0, column=0)
        button2.grid(row=0, column=2)
        button3.grid(row=6,column=2)
        window.geometry("350x150+800+1440")  # 主界面窗口显示
        window.title('披露易数据爬虫')
        window.mainloop()

    def getData(self):
        input_stock_code = self.input_stockCode()
        input_date = self.input_Date()
        self.get_stockdata(input_date, input_stock_code)
        self.file_save(participant_name_final=participant_name, shareholding_final=shareholding,
                          sharehoding_percent_final=sharehoding_percent, input_stock_code=input_stock_code,
                          input_date_data=input_date)

    def input_stockCode(self):  # 输入获取的股票代码

        while True:
            box = simpledialog.askstring(title='披露易数据爬取', prompt='请输入要爬取数据的股票代码：(格式：0XXXX)', initialvalue='0')
            if (str(box) != None and len(box) == 5 and str(box).isdigit() == True):  # 保证用户输入内容非空长度5且都为数字
                stockcode = str(box)
                return stockcode
            elif (len(box) < 5 or len(box) > 5 or str(box).isdigit() != True):
                warning = simpledialog.messagebox.showerror(title='严重警报', message='股票代码格式有误,唔好乱鬼咁输啊！')

    def input_Date(self):
        while True:
            box = simpledialog.askstring(title='披露易数据爬取', prompt='请输入要爬取数据的日期：(格式：XXXX/XX/XX)', initialvalue='2022/')
            if (str(box) != None and len(box) == 10):  # 保证用户输入内容非空长度为10且无字母
                data_date = str(box)
                return data_date
            elif (len(box) < 10 or len(box) > 10 or str(box).isalpha() != True):
                warning = simpledialog.messagebox.showerror(title='严重警报', message='日期格式有误,唔好乱鬼咁输啊！')

    def input_Date_Balance1(self):
        while True:
            box = simpledialog.askstring(title='披露易数据爬取', prompt='请输入要爬取的差额数据起始日期：(格式：XXXX/XX/XX)', initialvalue='2022/')
            if (str(box) != None and len(box) == 10):  # 保证用户输入内容非空长度为10且无字母
                data_date = str(box)
                return data_date
            elif (len(box) < 10 or len(box) > 10 or str(box).isalpha() != True):
                warning = simpledialog.messagebox.showerror(title='严重警报', message='日期格式有误,唔好乱鬼咁输啊！')

    def input_Date_Balance2(self):
        while True:
            box = simpledialog.askstring(title='披露易数据爬取', prompt='请输入要爬取的差额数据截止日期：(格式：XXXX/XX/XX)', initialvalue='2022/')
            if (str(box) != None and len(box) == 10):  # 保证用户输入内容非空长度为10且无字母
                data_date = str(box)
                return data_date
            elif (len(box) < 10 or len(box) > 10 or str(box).isalpha() != True):
                warning = simpledialog.messagebox.showerror(title='严重警报', message='日期格式有误,唔好乱鬼咁输啊！')





    def getBalance(self):#传入按钮参数command
        start_date=self.input_Date_Balance1()
        final_date=self.input_Date_Balance2()
        input_code=self.input_stockCode()
        self.get_stockdata(start_date,input_code)#爬取初始日期数据
        self.get_stockdata1(final_date,input_code)#爬取截止日期数据
        self.file_save2(participant_name_final=participant_name, shareholding_final=shareholding,
               sharehoding_percent_final=sharehoding_percent, input_stock_code=input_code,
               input_date_start=start_date,input_date_final=final_date)
        self.file_save1(participant_name_final=participant_name1, shareholding_final=shareholding1,
                       sharehoding_percent_final=sharehoding_percent1, input_stock_code=input_code,
                       input_date_final=final_date,input_date_start=start_date)

        #self.BalanceCalculate(start_date,final_date,final_file)#读取文档并对比券商并计算差额



    def BalanceCalculate(self,start_date,end_date,src):
        pass
        workbook=xlrd.open_workbook(src)




        # stock_name_key=list(get_sheet.rows[0])#券商名
        # stock_amount_key=list(get_sheet.rows[1])#持股量
        # total_key=list(get_sheet.rows[5])#持股量总数
        # issued_share_key=list(get_sheet.rows[6])#已发行股份/权证/单位





if __name__ == "__main__":
    dataget = HKex_Search()  # 创建对象
    dataget.main_window()  # 进入程序主窗口

'''1.返回当天披露易数据
   2.返回数据差额     '''
